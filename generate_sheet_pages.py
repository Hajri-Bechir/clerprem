import pandas as pd
import os
import re

# Get the absolute path to the project root directory
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(PROJECT_ROOT, 'data', 'Data.xlsx')
PAGES_DIR = 'pages'

# Ensure the 'pages' directory exists
if not os.path.exists(PAGES_DIR):
    os.makedirs(PAGES_DIR)
    print(f"Created directory: {PAGES_DIR}")

PAGE_TEMPLATE = """
import streamlit as st
import pandas as pd
import os
from openpyxl import load_workbook

# This page is for sheet: {sheet_name_original}
SHEET_NAME = "{sheet_name_original}"
# Use absolute path instead of relative path
EXCEL_FILE_PATH_IN_PAGE = r"{excel_file_absolute_path}" # Absolute path to Data.xlsx

st.set_page_config(page_title=f"Input: {sheet_name_title}", page_icon="📝")
st.title(f"📝 Data Input for Sheet: {{SHEET_NAME}}")

if not os.path.exists(EXCEL_FILE_PATH_IN_PAGE):
    st.error(f"Excel file not found at: {{EXCEL_FILE_PATH_IN_PAGE}}. Ensure 'data/Data.xlsx' exists.")
    st.stop()

try:
    df_sheet_structure = pd.read_excel(EXCEL_FILE_PATH_IN_PAGE, sheet_name=SHEET_NAME)
    columns = df_sheet_structure.columns.tolist()

    if not columns:
        st.warning(f"Sheet '{{SHEET_NAME}}' has no columns defined.")
        st.stop()

    with st.form(key=f"data_input_form_{{SHEET_NAME.replace(' ', '_')}}"):
        st.markdown("**Enter new data for the row:**")
        new_row_data = {{}}
        for col in columns:
            col_dtype = df_sheet_structure[col].dtype
            # Use a unique key for each widget by incorporating sheet name and column name
            widget_key = f"{{col.replace(' ', '_')}}_{{SHEET_NAME.replace(' ', '_')}}"
            if pd.api.types.is_numeric_dtype(col_dtype):
                new_row_data[col] = st.number_input(f"{{col}} (Numeric)", key=widget_key)
            elif pd.api.types.is_datetime64_any_dtype(col_dtype):
                # Fetching the current value if exists for date_input default (optional)
                # For a new row, this might not be applicable unless you want to default to last entry's date
                new_row_data[col] = st.date_input(f"{{col}} (Date)", value=None, key=widget_key)
            else:
                new_row_data[col] = st.text_input(f"{{col}} (Text)", key=widget_key)
        
        submit_button = st.form_submit_button(label=f"Add Row to Sheet: {{SHEET_NAME}}")

    if submit_button:
        try:
            # It's safer to read the existing workbook, append, then save.
            book = load_workbook(EXCEL_FILE_PATH_IN_PAGE)
            # Get the sheet, or create if it doesn't exist (though it should from initial read)
            sheet = book[SHEET_NAME] if SHEET_NAME in book.sheetnames else book.create_sheet(SHEET_NAME)
            
            # Create a new row from the dictionary values
            row_values = [new_row_data.get(col_name, None) for col_name in columns]
            sheet.append(row_values)
            book.save(EXCEL_FILE_PATH_IN_PAGE)

            st.success(f"Data successfully added to sheet '{{SHEET_NAME}}' and saved!")
            st.balloons()

        except Exception as e:
            st.error(f"Error saving data to Excel for sheet '{{SHEET_NAME}}': {{e}}")

    st.sidebar.markdown("--- ")
    if st.sidebar.checkbox(f"Show existing data for {{SHEET_NAME}}?", key=f"show_data_{{SHEET_NAME.replace(' ', '_')}}"):
        try:
            df_display = pd.read_excel(EXCEL_FILE_PATH_IN_PAGE, sheet_name=SHEET_NAME)
            st.sidebar.subheader(f"Current Data in '{{SHEET_NAME}}' (first 5 rows)")
            st.sidebar.dataframe(df_display.head())
        except Exception as e:
            st.sidebar.error(f"Could not display data for '{{SHEET_NAME}}': {{e}}")

except FileNotFoundError:
    st.error(f"ERROR: Excel file not found at {{EXCEL_FILE_PATH_IN_PAGE}}. Please ensure it exists and the path is correct.")
except KeyError as e:
    st.error(f"ERROR: Sheet '{{SHEET_NAME}}' not found in the Excel file. {{e}}. Available sheets: {{pd.ExcelFile(EXCEL_FILE_PATH_IN_PAGE).sheet_names}}")
except Exception as e:
    st.error(f"An unexpected error occurred while processing sheet '{{SHEET_NAME}}': {{e}}")
    st.info("Please ensure the Excel file and sheet exist and are correctly formatted.")

"""

def sanitize_filename_component(name):
    # Remove special characters, replace spaces with underscores for filename parts
    name = str(name) # Ensure it's a string
    name = re.sub(r'[^a-zA-Z0-9_\s-]', '', name).strip() # Allow alphanumeric, underscore, hyphen, space
    name = re.sub(r'[-\s]+', '_', name) # Replace spaces/hyphens with underscore
    return name

def create_pages():
    if not os.path.exists(EXCEL_FILE_PATH):
        print(f"Error: Excel file not found at {EXCEL_FILE_PATH}. Please create it or check the path.")
        return

    try:
        excel_file = pd.ExcelFile(EXCEL_FILE_PATH)
        sheet_names = excel_file.sheet_names
    except Exception as e:
        print(f"Error reading Excel file or sheet names: {e}")
        return

    if not sheet_names:
        print("No sheets found in the Excel file.")
        return

    print(f"Found sheets: {sheet_names}")

    # Clear previously auto-generated sheet pages (those matching the pattern)
    # This helps remove pages for sheets that no longer exist in Excel.
    # The numeric prefix helps Streamlit order pages.
    page_counter = 2 # Start page numbering from 02 (01_Home_Dashboard.py is 01)

    # First, remove old auto-generated pages to prevent clutter or errors from deleted sheets
    for item in os.listdir(PAGES_DIR):
        if re.match(r"^\d{2}_Input_.+\.py$", item):
            try:
                os.remove(os.path.join(PAGES_DIR, item))
                print(f"Removed old page: {item}")
            except OSError as e:
                print(f"Error removing old page {item}: {e}")

    for sheet_name_original in sheet_names:
        # Sanitize for display title (readable)
        sheet_name_title = sheet_name_original.replace("_", " ").title()
        # Sanitize for use in Python filenames/variables (more restrictive)
        safe_sheet_name_for_file = sanitize_filename_component(sheet_name_original)

        page_filename = f"{page_counter:02d}_Input_{safe_sheet_name_for_file}.py"
        page_filepath = os.path.join(PAGES_DIR, page_filename)

        # sheet_name_original is the exact name to be used for pandas operations
        # sheet_name_title is for user-facing titles
        
        # Get the absolute path to the Excel file
        excel_file_absolute_path = os.path.join(PROJECT_ROOT, 'data', 'Data.xlsx')
        # Replace backslashes with forward slashes for reliability
        excel_file_absolute_path = excel_file_absolute_path.replace('\\', '/')
        
        page_content = PAGE_TEMPLATE.format(
            sheet_name_original=sheet_name_original, 
            sheet_name_title=sheet_name_title,
            excel_file_absolute_path=excel_file_absolute_path
        )

        try:
            with open(page_filepath, 'w', encoding='utf-8') as f:
                f.write(page_content)
            print(f"Successfully created/updated page: {page_filepath} for sheet '{sheet_name_original}'")
            page_counter += 1
        except IOError as e:
            print(f"Error writing file {page_filepath}: {e}")

    print("\nPage generation process complete.")
    print(f"Please STOP your Streamlit application (if running) and RESTART it with: streamlit run 01_Home_Dashboard.py")
    print("This will allow Streamlit to detect the new/updated pages.")
    print(f"Once you confirm the new individual sheet pages work, you can delete the generic 'pages/02_Sheet_Data_Input.py'.")

if __name__ == "__main__":
    create_pages()
